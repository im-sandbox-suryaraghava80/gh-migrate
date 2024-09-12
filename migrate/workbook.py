import os
import pandas as pd
from loguru import logger

import pytz
import datetime

import openpyxl
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.worksheet.table import TableStyleInfo

# Create a table style
table_style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False
)


def initialize_workbook():
    workbook = load_workbook(os.path.join("report", "template", "workbook.xlsx"))
    workbook.save(os.path.join("report", "InfoMagnus - Migration Workbook.xlsx"))


def get_workbook(workbook_path):
    workbook = load_workbook(workbook_path)
    workbook.filename = workbook_path
    logger.info(f"** Found workbook at: {workbook_path}")

    return workbook


def get_mannequin_df(workbook_path):
    # Load the Org Mappings
    wb = load_workbook(workbook_path, data_only=True)

    ws = wb["Mapping - User"]

    data = list(ws.values)

    # Set the first row as the header
    df = pd.DataFrame(data[1:], columns=data[0])

    # Get orgs for wave, filter out excluded orgs
    users = df[(df["exclude"] == False)]

    # If orgs is empty
    if users.empty:
        raise ValueError("No users found in 'Mapping - User'")

    return users


def get_orgs_for_wave(org_type, wave, workbook_path):

    orgs = get_orgs_for_wave_df(wave, workbook_path)
    orgs = orgs[org_type].tolist()

    if orgs == ():
        raise ValueError("No {org_type} orgs found in 'Mapping - Org'")

    return orgs


def get_orgs_for_wave_df(wave, workbook_path):
    # Load the Org Mappings
    wb = load_workbook(workbook_path, data_only=True)

    ws = wb["Mapping - Org"]

    data = list(ws.values)

    # Set the first row as the header
    df = pd.DataFrame(data[1:], columns=data[0])

    # Get orgs for wave, filter out excluded orgs
    orgs = df[(df["exclude"] == False) & (df["wave"] == wave)]

    if orgs.empty:
        raise ValueError("No source orgs found in 'Mapping - Org'")

    return orgs


def autosize_columns(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column = column[0].column_letter
        for cell in worksheet[column]:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        adjusted_width = min(adjusted_width, 60)
        worksheet.column_dimensions[column].width = adjusted_width


def write_table(worksheet, df, table_name, heading=""):
    # If pushedAt exists in the df
    if "pushedAt" in df.columns:
        df["pushedAt"] = df["pushedAt"].dt.tz_localize(None)
    if "updatedAt" in df.columns:
        df["updatedAt"] = df["updatedAt"].dt.tz_localize(None)

    # Add a header with the table name
    if heading != "":
        worksheet.append([heading])
        worksheet[f"A{worksheet.max_row}"].font = Font(bold=True, size=12)

    # If there are no rows, add a single row with "No data"
    if df.empty:
        worksheet.append(["None"])
        worksheet[f"A{worksheet.max_row}"].font = Font(italic=True)
    else:
        # Add the table
        from openpyxl.worksheet.table import Table, TableStyleInfo

        num_rows, num_cols = df.shape
        if heading != "":
            start_col, start_row = "A", worksheet.max_row + 1
        else:
            start_col, start_row = "A", worksheet.max_row
        end_col = openpyxl.utils.get_column_letter(num_cols)
        end_row = start_row + num_rows

        table = Table(
            displayName=table_name,
            ref=f"{start_col}{start_row}:{end_col}{end_row}",
            tableStyleInfo=table_style,
        )
        worksheet.add_table(table)

        # Add the data
        worksheet.append(df.columns.to_list())
        for row in df.itertuples(index=False, name=None):
            worksheet.append(row)

        # Group added rows
        worksheet.row_dimensions.group(start_row + 1, end_row, hidden=False)
        autosize_columns(worksheet)

    # Move to the next empty row
    worksheet.append([])


def delete_worksheet(workbook, sheet_name):
    """ """
    # Delete the worksheet if it already exists
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    return workbook


def add_sheet(workbook, sheet_name, wave, desired_index=0, tab_color=None, delete=True):
    if not tab_color:
        if wave == 0:
            tab_color = "92D050"
        elif wave == 1:
            tab_color = "DAE9F8"
        elif wave == 2:
            tab_color = "A6C9EC"
        elif wave == 3:
            tab_color = "4D93D9"
        elif wave == 4:
            tab_color = "215C98"
        elif wave == 5:
            tab_color = "153D64"

    return _add_sheet(workbook, sheet_name, desired_index, tab_color, delete)


def _add_sheet(workbook, sheet_name, desired_index, tab_color, delete):
    """ """

    if delete:
        workbook = delete_worksheet(workbook, sheet_name)

    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name, index=desired_index)
        worksheet.sheet_properties.tabColor = tab_color

    return worksheet


def update_manns_worksheet(workbook, sheet_name, users):
    """ """
    desired_index = workbook.sheetnames.index("Cover") + 3
    worksheet = add_sheet(workbook, sheet_name, 0, desired_index, "002060")
    write_table(worksheet, stats, "User Mappings")


def add_inventory_worksheet(workbook, sheet_name, stats):
    """ """
    desired_index = workbook.sheetnames.index("Cover") + 3
    worksheet = add_sheet(workbook, sheet_name, 0, desired_index, "002060")
    write_table(worksheet, stats, "Inventory")


def write_mappings_file(df, cols):
    df = df[cols]

    # Rename the "name" columns to "source name" and "target name"
    cols = list(df.columns)

    # Rename the "name" columns to "source name" and "target name"
    if cols[0] == "name" and cols[1] == "name":
        cols[0] = "source_name"
        cols[1] = "target_name"
        cols[2] = "dry_run_target_name"
        cols[3] = "final_source_name"
        cols[4] = "final_target_name"
    else:
        raise ValueError('Columns 0 and 1 must be "name"')

    df.columns = cols
    return df


def add_org_mapping(workbook, sheet_name, stats):
    """ """

    desired_index = workbook.sheetnames.index("Cover") + 2
    worksheet = add_sheet(workbook, sheet_name, 0, desired_index, "FFC000")

    # TODO: Clean up this mess...
    stats = stats.rename(columns={"name": "_name"})
    stats = stats.rename(columns={"owner.login": "name"})
    stats["wave"] = 0
    stats["order"] = 0
    stats["exclude"] = False
    stats["exclude_reason"] = ""

    # Remove dupes from stats
    stats = stats.drop_duplicates(subset=["name"])

    stats = write_mappings_file(
        stats,
        [
            "name",
            "name",
            "name",
            "name",
            "name",
            "wave",
            "order",
            "exclude",
            "exclude_reason",
        ],
    )

    # Create org mapping table
    write_table(worksheet, stats, "Mapping_Org")

    workbook.save(workbook.filename)


def add_user_mapping(workbook, sheet_name, stats):
    """ """

    desired_index = workbook.sheetnames.index("Cover") + 2
    worksheet = add_sheet(workbook, sheet_name, 0, desired_index, "FFC000")

    workbook.save(workbook.filename)


def add_post_migration_timings_report(
    dry_run, wave, workbook, sheet_name, org_timings, repo_timings
):
    """ """
    desired_index = workbook.sheetnames.index("Cover") + 1
    worksheet = add_sheet(workbook, sheet_name, wave, desired_index)

    suffix = (
        f"migration_logs_dry_run_{wave}" if dry_run else f"migration_logs_prod_{wave}"
    )

    write_table(worksheet, org_timings, f"org_timings_{suffix}", "Org Timings")
    write_table(worksheet, repo_timings, f"repo_timings_{suffix}", "Repo Timings")

    workbook.save(workbook.filename)


def add_post_migration_logs_report(dry_run, wave, workbook, sheet_name, repo_results):
    """ """

    desired_index = workbook.sheetnames.index("Cover") + 1
    worksheet = add_sheet(workbook, sheet_name, wave, desired_index)

    # Clear the contents of the worksheet
    worksheet.delete_rows(1, worksheet.max_row)

    suffix = (
        f"migration_logs_dry_run_{wave}" if dry_run else f"migration_logs_prod_{wave}"
    )

    write_table(worksheet, repo_results, f"repo_results_{suffix}", "Repo Warnings")

    workbook.save(workbook.filename)


def add_post_migration_stats_report(dry_run, wave, workbook, sheet_name, stats):
    """ """

    desired_index = workbook.sheetnames.index("Cover") + 1
    worksheet = add_sheet(workbook, sheet_name, wave, desired_index)

    # Clear the contents of the worksheet
    worksheet.delete_rows(1, worksheet.max_row)

    suffix = f"_stats_dry_run_{wave}" if dry_run else f"_stats_prod_{wave}"
    write_table(worksheet, stats, f"Stats{suffix}", "Post-Migration Stats Report")


def add_post_migration_snaps_report(
    dry_run, wave, workbook, sheet_name, repos, users, teams, team_users, team_repos
):
    """ """

    desired_index = workbook.sheetnames.index("Cover") + 1
    worksheet = add_sheet(workbook, sheet_name, wave, desired_index)

    suffix = f"snapshots_dry_run_{wave}" if dry_run else f"snapshots_prod_{wave}"

    write_table(worksheet, repos, f"repos_{suffix}", "Repos")
    write_table(worksheet, users, f"users_{suffix}", "Users")
    write_table(worksheet, teams, f"teams_{suffix}", "Teams")
    write_table(worksheet, team_users, f"team_users_{suffix}", "Team Users")
    write_table(worksheet, team_repos, f"team_repos_{suffix}", "Team Repos")

    workbook.save(workbook.filename)


def add_pre_migration_report(workbook, sheet_name, stats):
    """ """
    desired_index = workbook.sheetnames.index("Cover") + 1
    worksheet = add_sheet(workbook, sheet_name, 0, desired_index, "7030A0")

    # Create large repos table
    df = stats[stats["diskUsage"] > 1000000].sort_values(by="diskUsage")
    write_table(worksheet, df, "Large_Repos", "Large Repos")

    # Create large PRs table
    df = stats[stats["pullRequests.totalCount"] > 1000].sort_values(
        by="pullRequests.totalCount"
    )
    write_table(worksheet, df, "Large_PRs", "Large PRs")

    # Create webhooks table
    df = stats[stats["webhooks.totalCount"] > 0].sort_values(by="webhooks.totalCount")
    write_table(worksheet, df, "Webhooks_Repos", "Repos with webhooks")

    # Create actions table
    df = stats[stats["lastWorkflowRun"] != None].sort_values("lastWorkflowRun")
    write_table(worksheet, df, "Actions_Repos", "Repos with actions")

    # Create stale repos table
    stats["pushedAt"] = stats["pushedAt"].dt.tz_localize("UTC")
    df = stats[
        stats["pushedAt"]
        < datetime.datetime.now(pytz.UTC) - datetime.timedelta(days=60)
    ].sort_values("pushedAt")
    write_table(worksheet, df, "Stale_Repos", "Stale Repos")

    # Create archived repos table
    df = stats[stats["isArchived"] == True].sort_values("isArchived")
    write_table(worksheet, df, "Archived_Repos", "Archived Repos")

    # Create locked repos table
    df = stats[stats["isLocked"] == True].sort_values("isLocked")
    write_table(worksheet, df, "Locked_Repos", "Locked Repos")

    # Create repos with packages table
    df = stats[stats["packages.totalCount"] > 0].sort_values("packages.totalCount")
    write_table(worksheet, df, "Has_Packages", "Repos with packages")

    # def identify_git_lfs():
    # # TODO: Need to figure out how to implement this

    workbook.save(workbook.filename)
